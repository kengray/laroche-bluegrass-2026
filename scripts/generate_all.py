# =============================================================================
# generate_all.py
# Reads data/schedule.json and generates:
#   - LaRoche2026.ics  (subscribable calendar file)
#   - LaRoche2026_Festival_Schedule.xlsx  (formatted spreadsheet)
#
# Run locally:  python generate_all.py
# In CI:        triggered automatically by GitHub Actions on push to schedule.json
#
# Requires: openpyxl  (pip install openpyxl)
# =============================================================================

import json
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Load schedule data from JSON
# ---------------------------------------------------------------------------
data_path = Path(__file__).parent.parent / "data" / "schedule.json"
# In the repo: script is at scripts/generate_all.py, data at data/schedule.json
# parent.parent goes from scripts/ up to repo root
# Fallback for local testing outside repo structure:
if not data_path.exists():
    data_path = Path(__file__).parent / "data" / "schedule.json"
with open(data_path, encoding="utf-8") as f:
    data = json.load(f)

FESTIVAL_LOCATION  = data["festival"]["location"]
CAMP_LOCATION      = data["festival"]["camp_location"]
FESTIVAL_NAME      = data["festival"]["name"]

# ---------------------------------------------------------------------------
# Date lookup: display string -> DD/MM/YYYY
# ---------------------------------------------------------------------------
DATE_MAP = {
    "Mon 27 Jul": "27/07/2026",
    "Tue 28 Jul": "28/07/2026",
    "Wed 29 Jul": "29/07/2026",
    "Thu 30 Jul": "30/07/2026",
    "Fri 31 Jul": "31/07/2026",
    "Sat 1 Aug":  "01/08/2026",
    "Sun 2 Aug":  "02/08/2026",
}

# =============================================================================
# ICS GENERATION
# =============================================================================

def make_ics_event(uid, summary, date_str, start_time, end_time, description, location):
    """Build a single VEVENT block. Handles acts that finish past midnight."""
    date     = datetime.strptime(date_str, "%d/%m/%Y")
    sh, sm   = int(start_time[:2]), int(start_time[3:])
    eh, em   = int(end_time[:2]),   int(end_time[3:])
    start_dt = date.replace(hour=sh, minute=sm)
    end_dt   = date.replace(hour=eh, minute=em)
    if end_dt <= start_dt:          # crosses midnight
        end_dt += timedelta(days=1)
    fmt  = "%Y%m%dT%H%M%S"
    desc = description.replace(",", "\\,").replace("\n", "\\n")
    summ = summary.replace(",", "\\,")
    return (
        f"BEGIN:VEVENT\n"
        f"UID:{uid}@larochebluegrass2026\n"
        f"DTSTART;TZID=Europe/Paris:{start_dt.strftime(fmt)}\n"
        f"DTEND;TZID=Europe/Paris:{end_dt.strftime(fmt)}\n"
        f"SUMMARY:{summ}\n"
        f"DESCRIPTION:{desc}\n"
        f"LOCATION:{location}\n"
        f"END:VEVENT"
    )

festival_events = []
camp_events     = []
counter = 1

# Main Stage
for row in data["main_stage"]:
    summary = f"🎸 {row['name']} — Main Stage | {FESTIVAL_NAME}"
    festival_events.append(make_ics_event(
        f"blr2026-{counter:03d}", summary,
        DATE_MAP[row["date"]], row["start"], row["end"],
        row["notes"], FESTIVAL_LOCATION
    ))
    counter += 1

# Day Stage
for row in data["day_stage"]:
    summary = f"🎸 {row['name']} — Day Stage | {FESTIVAL_NAME}"
    festival_events.append(make_ics_event(
        f"blr2026-{counter:03d}", summary,
        DATE_MAP[row["date"]], row["start"], row["end"],
        row["notes"], FESTIVAL_LOCATION
    ))
    counter += 1

# Street Festival
for row in data["street_festival"]:
    stage   = row.get("stage", "Street Festival")
    summary = f"🎸 {row['name']} — {stage} | {FESTIVAL_NAME}"
    start   = row["start"] if row["start"] else "19:00"
    end     = row["end"]   if row["end"]   else "23:00"
    festival_events.append(make_ics_event(
        f"blr2026-{counter:03d}", summary,
        DATE_MAP[row["date"]], start, end,
        row["notes"], FESTIVAL_LOCATION
    ))
    counter += 1

# Teaching Camp
for row in data["teaching_camp"]:
    summary = f"🎓 {row['name']} — Teaching Camp | {FESTIVAL_NAME}"
    camp_events.append(make_ics_event(
        f"blr2026-{counter:03d}", summary,
        DATE_MAP[row["date"]], row["start"], row["end"],
        row["notes"], CAMP_LOCATION
    ))
    counter += 1

all_events = festival_events + camp_events

# Output files always go to the repo root (one level up from scripts/)
repo_root = Path(__file__).parent.parent
# Fallback for local testing
if not (repo_root / "data").exists():
    repo_root = Path(__file__).parent

xlsx_path = repo_root / "LaRoche2026_Festival_Schedule.xlsx"


def write_ics(path, cal_name, cal_desc, events_list):
    content = (
        "BEGIN:VCALENDAR\n"
        "VERSION:2.0\n"
        f"PRODID:-//{FESTIVAL_NAME}//EN\n"
        "CALSCALE:GREGORIAN\n"
        "METHOD:PUBLISH\n"
        f"X-WR-CALNAME:{cal_name}\n"
        "X-WR-TIMEZONE:Europe/Paris\n"
        f"X-WR-CALDESC:{cal_desc}\n"
        + "\n".join(events_list)
        + "\nEND:VCALENDAR\n"
    )
    Path(path).write_text(content, encoding="utf-8")
    print(f"ICS written: {path.name} ({len(events_list)} events)")


write_ics(
    repo_root / "LaRoche2026-Festival.ics",
    f"{FESTIVAL_NAME} — Concerts",
    f"Main Stage\\, Day Stage and Street Festival concerts. {FESTIVAL_NAME}\\, La Roche-sur-Foron\\, France. 30 July - 2 August 2026.",
    festival_events
)
write_ics(
    repo_root / "LaRoche2026-Camp.ics",
    f"{FESTIVAL_NAME} — Teaching Camp",
    f"Adult teaching camp timetable. Lycée Sainte Famille\\, La Roche-sur-Foron\\, France. 27-30 July 2026.",
    camp_events
)
write_ics(
    repo_root / "LaRoche2026-Full.ics",
    f"{FESTIVAL_NAME} — Full Schedule",
    f"Complete schedule including concerts and teaching camp. {FESTIVAL_NAME}\\, La Roche-sur-Foron\\, France.",
    all_events
)
# Legacy file — keeps existing subscribers updated
write_ics(
    repo_root / "LaRoche2026.ics",
    f"{FESTIVAL_NAME} — Full Schedule",
    f"Complete schedule including concerts and teaching camp. {FESTIVAL_NAME}\\, La Roche-sur-Foron\\, France.",
    all_events
)


# =============================================================================
# XLSX GENERATION
# =============================================================================

# Colour palette
DARK_GREEN   = "1A3C34"
MID_GREEN    = "2D6A4F"
LIGHT_GREEN  = "D8F3DC"
GOLD         = "B7950B"
GOLD_LIGHT   = "FFF9C4"
WHITE        = "FFFFFF"
GREY_ROW     = "F4F4F4"
DARK_RED     = "8B0000"
STREET_BLUE  = "1A3A5C"
STREET_LIGHT = "D6E4F0"
CAMP_PURPLE  = "4A235A"
CAMP_LIGHT   = "EAD5F5"


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def make_sheet(wb, title, rows, tab_color, header_color, light_row_color, is_first=False):
    """
    Build a formatted worksheet from a list of row dicts.
    Each row dict must have: date, name, country, start, end, notes.
    Optionally: stage (used as Stage Note column).
    """
    ws = wb.active if is_first else wb.create_sheet(title)
    ws.title = title
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color

    headers    = ["Date", "Band / Act", "Country", "Stage Note", "Start Time", "End Time", "Notes"]
    col_widths = [14,      36,           22,         20,           13,           13,          45]

    # Title banner
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = f"{FESTIVAL_NAME} — {title}"
    c.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=DARK_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells("A2:G2")
    c           = ws["A2"]
    c.value     = "Source data: data/schedule.json in the GitHub repository"
    c.font      = Font(name="Arial", italic=True, size=9, color="666666")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Column headers
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=3, column=col, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20

    # Example row
    example = ["e.g. Thu 30 Jul", "e.g. Kristy Cox Band", "AU/US", "Main Stage",
               "19:00", "20:00", "Times in 24hr format"]
    for col, val in enumerate(example, 1):
        cell           = ws.cell(row=4, column=col, value=val)
        cell.font      = Font(name="Arial", size=9, italic=True, color="AAAAAA")
        cell.fill      = PatternFill("solid", fgColor="F0F0F0")
        cell.border    = thin_border()
        cell.alignment = Alignment(
            vertical="center",
            horizontal="center" if col in (5, 6) else "left",
            wrap_text=True
        )
    ws.row_dimensions[4].height = 24

    # Data rows
    current_date = None
    for i, row in enumerate(rows, start=5):
        is_tbc = row["name"].startswith("TBC")
        if row["date"] != current_date:
            current_date = row["date"]
            bg = light_row_color
        else:
            bg = GREY_ROW if i % 2 == 0 else WHITE
        if is_tbc:
            bg = "FFF3CD"

        values = [
            row["date"],
            row["name"],
            row.get("country", row.get("location", "")),
            row.get("stage", title),
            row["start"],
            row["end"],
            row["notes"],
        ]
        for col, val in enumerate(values, 1):
            cell           = ws.cell(row=i, column=col, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font      = Font(name="Arial", size=9, italic=is_tbc,
                                  color="888888" if is_tbc else "000000")
            if col in (5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Arial", size=9, color=GOLD,
                                 bold=not is_tbc, italic=is_tbc)
        ws.row_dimensions[i].height = 30

    ws.freeze_panes = "A5"
    return ws


wb = Workbook()
make_sheet(wb, "Main Stage",      data["main_stage"],      MID_GREEN,   MID_GREEN,   LIGHT_GREEN, is_first=True)
make_sheet(wb, "Day Stage",       data["day_stage"],       GOLD,        GOLD,        GOLD_LIGHT,  is_first=False)
make_sheet(wb, "Street Festival", data["street_festival"], STREET_BLUE, STREET_BLUE, STREET_LIGHT,is_first=False)

# Teaching camp rows need a "stage" key for the Stage Note column
camp_rows = []
for row in data["teaching_camp"]:
    r = dict(row)
    r["country"] = r.pop("location", "")
    r["stage"]   = "Teaching Camp"
    camp_rows.append(r)
make_sheet(wb, "Teaching Camp", camp_rows, CAMP_PURPLE, CAMP_PURPLE, CAMP_LIGHT, is_first=False)

wb.save(xlsx_path)
print(f"XLSX written: {xlsx_path}")


# =============================================================================
# HTML GENERATION
# Produces index.html — the GitHub Pages band guide webpage.
# All band data comes from schedule.json so a single edit updates everything.
# =============================================================================

def js_str(s):
    """Escape a string for safe embedding in a JS string literal."""
    return s.replace("\\", "\\\\").replace('"', '\\"').replace("\n", " ").replace("'", "\\'")

def build_band_js(band, stage_label):
    """Convert a schedule.json band dict into a JS object literal string."""
    name     = js_str(band.get("name", ""))
    country  = js_str(band.get("country", ""))
    notes    = js_str(band.get("notes", ""))
    notes_fr = js_str(band.get("notes_fr", "") or band.get("notes", ""))
    date     = js_str(band.get("date", ""))
    start    = js_str(band.get("start", ""))
    website  = js_str(band.get("website", "") or "")
    video    = js_str(band.get("video", "") or "")
    spotify  = js_str(band.get("spotify", "") or "")
    return (
        f'  {{ date:"{date}", name:"{name}", country:"{country}", '
        f'stage:"{stage_label}", time:"{start}", notes:"{notes}", notes_fr:"{notes_fr}", '
        f'website:"{website}", video:"{video}", spotify:"{spotify}" }}'
    )

# Build the JS bands array from all sections
js_bands = []
for b in data["main_stage"]:
    b2 = dict(b)
    js_bands.append(build_band_js(b2, "Main Stage"))
for b in data["day_stage"]:
    b2 = dict(b)
    js_bands.append(build_band_js(b2, "Day Stage"))
# Street festival excluded from band guide (TBC placeholders not useful here)

bands_js = "[\n" + ",\n".join(js_bands) + "\n]"

html_content = f'''<!DOCTYPE html>
<html lang="en" id="html-root">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{FESTIVAL_NAME}</title>
<link rel="icon" type="image/svg+xml" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><rect width='100' height='100' rx='20' fill='%231a3c34'/><text y='72' x='50' text-anchor='middle' font-size='62'>🪕</text></svg>">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;600&family=Source+Sans+3:wght@300;400;500&display=swap" rel="stylesheet">
<style>
  :root {{
    --green-dark: #1a3c34;
    --green-mid: #2d6a4f;
    --green-light: #d8f3dc;
    --gold: #b7950b;
    --gold-light: #f9f3d0;
    --cream: #faf8f3;
    --text: #1a1a18;
    --text-muted: #5a5a54;
    --border: #e0ddd5;
    --white: #ffffff;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: "Source Sans 3", sans-serif; background: var(--cream); color: var(--text); font-size: 16px; line-height: 1.6; }}
  .hero {{ background: #111827; color: var(--white); padding: 6rem 2rem 5rem; text-align: center; position: relative; overflow: hidden; }}
  .hero-poster {{ position: absolute; inset: 0; width: 100%; height: 100%; object-fit: cover; object-position: center 40%; opacity: 0.75; z-index: 0; }}
  .hero::after {{ content: ""; position: absolute; inset: 0; background: linear-gradient(to bottom, rgba(17,24,39,0.45) 0%, rgba(17,24,39,0.65) 100%); z-index: 1; }}
  .hero-inner {{ position: relative; z-index: 2; max-width: 680px; margin: 0 auto; text-shadow: 0 1px 8px rgba(0,0,0,0.6); }}
  .hero-label {{ font-size: 11px; font-weight: 500; letter-spacing: 0.2em; text-transform: uppercase; color: rgba(255,255,255,0.75); margin-bottom: 1rem; }}
  .hero h1 {{ font-family: "Playfair Display", serif; font-size: clamp(2rem, 6vw, 3.2rem); font-weight: 600; line-height: 1.15; margin-bottom: 0.75rem; }}
  .hero-sub {{ font-size: 16px; color: rgba(255,255,255,0.65); margin-bottom: 2rem; }}
  .subscribe-btn {{ display: inline-flex; align-items: center; gap: 8px; background: var(--gold); color: var(--white); padding: 12px 28px; border-radius: 4px; text-decoration: none; font-size: 14px; font-weight: 500; letter-spacing: 0.03em; transition: background 0.15s; }}
  .subscribe-btn:hover {{ background: #9a7d09; }}
  .subscribe-btn svg {{ width: 16px; height: 16px; }}
  .content {{ max-width: 1100px; margin: 0 auto; padding: 2.5rem 1.5rem; }}
  .filters {{ display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 2.5rem; padding-bottom: 1.5rem; border-bottom: 1px solid var(--border); }}
  .filter-btn {{ font-family: "Source Sans 3", sans-serif; font-size: 13px; font-weight: 400; padding: 6px 16px; border-radius: 20px; border: 1px solid var(--border); background: var(--white); color: var(--text-muted); cursor: pointer; transition: all 0.15s; }}
  .filter-btn:hover {{ border-color: var(--green-mid); color: var(--green-mid); }}
  .filter-btn.active {{ background: var(--green-dark); border-color: var(--green-dark); color: var(--white); }}
  .day-section {{ margin-bottom: 3rem; }}
  .day-header {{ display: flex; align-items: baseline; gap: 12px; margin-bottom: 1.25rem; }}
  .day-title {{ font-family: "Playfair Display", serif; font-size: 20px; font-weight: 600; color: var(--green-dark); }}
  .day-count {{ font-size: 12px; color: var(--text-muted); font-weight: 300; }}
  .band-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(300px, 1fr)); gap: 1px; background: var(--border); border: 1px solid var(--border); border-radius: 8px; overflow: hidden; }}
  .band-card {{ background: var(--white); padding: 1.25rem 1.5rem; transition: background 0.1s; cursor: pointer; }}
  .band-card:hover {{ background: #fdfcf8; }}
  .band-card.expanded {{ background: #f7faf8; border-left: 3px solid var(--green-mid); }}
  .card-header {{ display: flex; justify-content: space-between; align-items: flex-start; gap: 8px; margin-bottom: 6px; }}
  .band-name {{ font-family: "Playfair Display", serif; font-size: 16px; font-weight: 600; color: var(--text); line-height: 1.3; }}
  .stage-pill {{ font-size: 10px; font-weight: 500; letter-spacing: 0.08em; text-transform: uppercase; padding: 3px 8px; border-radius: 3px; white-space: nowrap; flex-shrink: 0; margin-top: 2px; }}
  .pill-main {{ background: var(--green-light); color: var(--green-mid); }}
  .pill-day  {{ background: var(--gold-light); color: var(--gold); }}
  .card-meta {{ font-size: 12px; color: var(--text-muted); margin-bottom: 8px; font-weight: 300; }}
  .card-notes {{ font-size: 13px; color: var(--text-muted); line-height: 1.55; margin-bottom: 12px; }}
  .expand-ellipsis {{ display: inline; background: var(--cream); border: 1px solid var(--border); border-radius: 3px; padding: 0 4px; font-size: 12px; color: var(--green-mid); cursor: pointer; font-weight: 500; margin-left: 2px; }}
  .expand-ellipsis:hover {{ background: var(--green-light); border-color: var(--green-mid); }}
  .card-links {{ display: flex; gap: 6px; flex-wrap: wrap; }}
  .card-link {{ font-size: 11px; font-weight: 500; letter-spacing: 0.04em; text-transform: uppercase; padding: 4px 10px; border-radius: 3px; border: 1px solid var(--border); color: var(--text-muted); text-decoration: none; background: transparent; transition: all 0.12s; }}
  .card-link:hover {{ border-color: var(--green-mid); color: var(--green-dark); background: var(--green-light); }}
  .card-link.spotify:hover {{ border-color: #1db954; color: #1db954; background: #edfbf2; }}
  .subscribe-section {{ background: var(--white); border: 1px solid var(--border); border-radius: 8px; padding: 2rem; margin-bottom: 2.5rem; }}
  .subscribe-section-header {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 0.5rem; }}
  .subscribe-section h2 {{ font-family: "Playfair Display", serif; font-size: 20px; font-weight: 600; color: var(--green-dark); }}
  .close-btn {{ font-size: 18px; color: var(--text-muted); background: none; border: none; cursor: pointer; line-height: 1; padding: 0 4px; }}
  .close-btn:hover {{ color: var(--text); }}
  .subscribe-section > p {{ font-size: 14px; color: var(--text-muted); margin-bottom: 1.5rem; }}
  .platform-tabs {{ display: flex; gap: 6px; margin-bottom: 1.5rem; flex-wrap: wrap; }}
  .tab-btn {{ font-size: 13px; padding: 6px 16px; border-radius: 20px; border: 1px solid var(--border); background: transparent; color: var(--text-muted); cursor: pointer; transition: all 0.15s; }}
  .tab-btn.active {{ background: var(--green-dark); border-color: var(--green-dark); color: var(--white); }}
  .tab-content {{ display: none; }}
  .tab-content.active {{ display: block; }}
  .steps ol {{ padding-left: 1.25rem; }}
  .steps li {{ font-size: 14px; color: var(--text); line-height: 1.6; margin-bottom: 0.6rem; }}
  .steps li span {{ color: var(--text-muted); }}
  .cal-url {{ display: flex; align-items: center; gap: 8px; background: var(--cream); border: 1px solid var(--border); border-radius: 4px; padding: 8px 12px; margin: 0.75rem 0; font-family: monospace; font-size: 13px; color: var(--text); word-break: break-all; }}
  .copy-btn {{ font-size: 11px; font-weight: 500; padding: 4px 10px; border-radius: 3px; border: 1px solid var(--border); background: var(--white); color: var(--text-muted); cursor: pointer; white-space: nowrap; flex-shrink: 0; transition: all 0.12s; }}
  .copy-btn:hover {{ border-color: var(--green-mid); color: var(--green-dark); }}
  .note {{ font-size: 13px; color: var(--text-muted); background: var(--gold-light); border-left: 3px solid var(--gold); padding: 8px 12px; border-radius: 0 4px 4px 0; margin-top: 1rem; }}
  footer {{ text-align: center; padding: 2rem; font-size: 12px; color: var(--text-muted); border-top: 1px solid var(--border); margin-top: 2rem; }}
  footer a {{ color: var(--green-mid); text-decoration: none; }}
  footer a:hover {{ text-decoration: underline; }}
  .lang-toggle {{ position: absolute; top: 1rem; right: 1.5rem; z-index: 3; display: flex; gap: 6px; }}
  .lang-btn {{ background: rgba(255,255,255,0.15); border: 1px solid rgba(255,255,255,0.3); border-radius: 4px; padding: 4px 8px; cursor: pointer; font-size: 20px; line-height: 1; transition: background 0.15s; }}
  .lang-btn:hover {{ background: rgba(255,255,255,0.3); }}
  .lang-btn.active {{ background: rgba(255,255,255,0.35); border-color: rgba(255,255,255,0.7); }}
  [data-en] {{ display: block; }}
  [data-fr] {{ display: none; }}
  .fr [data-en] {{ display: none; }}
  .fr [data-fr] {{ display: block; }}
  span[data-en], span[data-fr] {{ display: inline; }}
  .fr span[data-en] {{ display: none; }}
  .fr span[data-fr] {{ display: inline; }}
  @media (max-width: 600px) {{ .hero {{ padding: 3rem 1.25rem 2rem; }} .content {{ padding: 1.5rem 1rem; }} .band-grid {{ grid-template-columns: 1fr; }} .lang-toggle {{ top: 0.75rem; right: 0.75rem; }} }}
</style>
</head>
<body>
<header class="hero">
  <img src="poster.jpg" alt="" class="hero-poster">
  <div class="lang-toggle">
    <button class="lang-btn active" onclick="setLang('en',this)" title="English">🏴󠁧󠁢󠁳󠁣󠁴󠁿</button>
    <button class="lang-btn" onclick="setLang('fr',this)" title="Français">🇫🇷</button>
  </div>
  <div class="hero-inner">
    <p class="hero-label">La Roche-sur-Foron, France</p>
    <h1>{FESTIVAL_NAME}</h1>
    <p class="hero-sub">30 July &ndash; 2 August &nbsp;&middot;&nbsp; <span data-en>28 concerts &nbsp;&middot;&nbsp; 22 bands &nbsp;&middot;&nbsp; 19 countries</span><span data-fr>28 concerts &nbsp;&middot;&nbsp; 22 groupes &nbsp;&middot;&nbsp; 19 pays</span></p>
    <a class="subscribe-btn" href="#" onclick="toggleInstructions(event)">
      <svg viewBox="0 0 16 16" fill="none" stroke="currentColor" stroke-width="1.5"><rect x="1" y="3" width="14" height="11" rx="1.5"/><path d="M1 6h14M5 1v4M11 1v4"/></svg>
      <span data-en>Subscribe to calendar</span><span data-fr>S'abonner au calendrier</span>
    </a>
  </div>
</header>
<main class="content">

  <div class="subscribe-section" id="subscribe-section" style="display:none;">
    <div class="subscribe-section-header">
      <h2><span data-en>Add to your calendar</span><span data-fr>Ajouter à votre calendrier</span></h2>
      <button class="close-btn" onclick="toggleInstructions(event)" title="Close">&times;</button>
    </div>
    <p><span data-en>Subscribe once and your calendar updates automatically whenever the schedule changes. Choose which calendar you want, then select your platform.</span><span data-fr>Abonnez-vous une seule fois et votre calendrier se met à jour automatiquement dès que le programme change. Choisissez le calendrier souhaité, puis sélectionnez votre plateforme.</span></p>
    <div class="platform-tabs" style="margin-bottom:1rem;">
      <button class="tab-btn active" onclick="setCalType('festival',this)"><span data-en>Concerts only</span><span data-fr>Concerts uniquement</span></button>
      <button class="tab-btn" onclick="setCalType('camp',this)"><span data-en>Teaching camp only</span><span data-fr>Stage uniquement</span></button>
      <button class="tab-btn" onclick="setCalType('full',this)"><span data-en>Full schedule</span><span data-fr>Programme complet</span></button>
    </div>
    <div class="cal-url" id="cal-url-display">https://tinyurl.com/LaRoche2026-Festival <button class="copy-btn" onclick="copyUrl()"><span data-en>Copy</span><span data-fr>Copier</span></button></div>
    <div class="platform-tabs">
      <button class="tab-btn active" onclick="showTab('android',this)">Android</button>
      <button class="tab-btn" onclick="showTab('iphone',this)">iPhone / iPad</button>
      <button class="tab-btn" onclick="showTab('google',this)"><span data-en>Google Calendar (desktop)</span><span data-fr>Google Agenda (ordinateur)</span></button>
      <button class="tab-btn" onclick="showTab('outlook',this)">Outlook</button>
      <button class="tab-btn" onclick="showTab('apple',this)"><span data-en>Apple Calendar (Mac)</span><span data-fr>Calendrier Apple (Mac)</span></button>
    </div>

    <div id="tab-android" class="tab-content active steps">
      <div data-en>
      <ol>
        <li>On your Android phone, open a browser and go to <strong>calendar.google.com</strong> — the app itself does not support adding subscriptions directly.</li>
        <li>Tap the menu icon (three lines, top left) and select <strong>Other calendars</strong>, then the <strong>+</strong> button.</li>
        <li>Choose <strong>From URL</strong>.</li>
        <li>Paste the calendar URL shown above, then tap <strong>Add calendar</strong>.</li>
        <li>Open the Google Calendar app — events appear within a few minutes. Future updates sync automatically, usually within 24 hours.</li>
      </ol>
      <p class="note">The Google Calendar app on Android does not have a "subscribe by URL" option — you must use the browser-based calendar.google.com to add it. This is a Google limitation, not a problem with the calendar itself.</p>
      </div>
      <div data-fr>
      <ol>
        <li>Sur votre téléphone Android, ouvrez un navigateur et allez sur <strong>calendar.google.com</strong> — l'application elle-même ne prend pas en charge l'ajout d'abonnements directement.</li>
        <li>Appuyez sur l'icône de menu (trois lignes, en haut à gauche) et sélectionnez <strong>Autres agendas</strong>, puis le bouton <strong>+</strong>.</li>
        <li>Choisissez <strong>À partir de l'URL</strong>.</li>
        <li>Collez l'URL du calendrier indiquée ci-dessus, puis appuyez sur <strong>Ajouter un agenda</strong>.</li>
        <li>Ouvrez l'application Google Agenda — les événements apparaîtront en quelques minutes. Les mises à jour futures se synchronisent automatiquement, généralement sous 24 heures.</li>
      </ol>
      <p class="note">L'application Google Agenda sur Android ne dispose pas d'option d'abonnement par URL — vous devez utiliser calendar.google.com dans un navigateur. Il s'agit d'une limitation de Google, pas d'un problème avec le calendrier.</p>
      </div>
    </div>

    <div id="tab-iphone" class="tab-content steps">
      <div data-en>
      <ol>
        <li>Open the <strong>Settings</strong> app.</li>
        <li>Scroll down and tap <strong>Calendar</strong>, then <strong>Accounts</strong>.</li>
        <li>Tap <strong>Add Account</strong>, then choose <strong>Other</strong>.</li>
        <li>Tap <strong>Add Subscribed Calendar</strong>.</li>
        <li>Paste the calendar URL shown above, then tap <strong>Next</strong> and <strong>Save</strong>.</li>
      </ol>
      <p class="note">Updates sync automatically. You can control the sync frequency in Settings &gt; Calendar &gt; Accounts &gt; Fetch New Data.</p>
      </div>
      <div data-fr>
      <ol>
        <li>Ouvrez l'application <strong>Réglages</strong>.</li>
        <li>Faites défiler vers le bas et appuyez sur <strong>Calendrier</strong>, puis <strong>Comptes</strong>.</li>
        <li>Appuyez sur <strong>Ajouter un compte</strong>, puis choisissez <strong>Autre</strong>.</li>
        <li>Appuyez sur <strong>Ajouter un calendrier avec abonnement</strong>.</li>
        <li>Collez l'URL indiquée ci-dessus, puis appuyez sur <strong>Suivant</strong> et <strong>Enregistrer</strong>.</li>
      </ol>
      <p class="note">Les mises à jour se synchronisent automatiquement. Vous pouvez contrôler la fréquence dans Réglages &gt; Calendrier &gt; Comptes &gt; Nouvelles données.</p>
      </div>
    </div>

    <div id="tab-google" class="tab-content steps">
      <div data-en>
      <ol>
        <li>Go to <strong>calendar.google.com</strong> in your browser.</li>
        <li>On the left sidebar, find <strong>Other calendars</strong> and click the <strong>+</strong> button next to it.</li>
        <li>Choose <strong>From URL</strong>.</li>
        <li>Paste the calendar URL shown above, then click <strong>Add calendar</strong>.</li>
      </ol>
      <p class="note">Google Calendar re-syncs subscribed calendars roughly every 24 hours, so updates may not appear immediately after a schedule change.</p>
      </div>
      <div data-fr>
      <ol>
        <li>Allez sur <strong>calendar.google.com</strong> dans votre navigateur.</li>
        <li>Dans la barre latérale gauche, trouvez <strong>Autres agendas</strong> et cliquez sur le bouton <strong>+</strong>.</li>
        <li>Choisissez <strong>À partir de l'URL</strong>.</li>
        <li>Collez l'URL indiquée ci-dessus, puis cliquez sur <strong>Ajouter un agenda</strong>.</li>
      </ol>
      <p class="note">Google Agenda resynchronise les agendas abonnés environ toutes les 24 heures, les mises à jour peuvent donc ne pas apparaître immédiatement.</p>
      </div>
    </div>

    <div id="tab-outlook" class="tab-content steps">
      <div data-en>
      <ol>
        <li>Open <strong>Outlook</strong> (desktop app or outlook.com).</li>
        <li>Go to the <strong>Calendar</strong> view.</li>
        <li>Click <strong>Add calendar</strong> (or <strong>Open calendar</strong> in the desktop app).</li>
        <li>Choose <strong>Subscribe from web</strong> (or <strong>From internet</strong> in the desktop app).</li>
        <li>Paste the calendar URL shown above, then click <strong>Import</strong> or <strong>OK</strong>.</li>
      </ol>
      </div>
      <div data-fr>
      <ol>
        <li>Ouvrez <strong>Outlook</strong> (application de bureau ou outlook.com).</li>
        <li>Allez dans la vue <strong>Calendrier</strong>.</li>
        <li>Cliquez sur <strong>Ajouter un calendrier</strong> (ou <strong>Ouvrir le calendrier</strong> dans l'application de bureau).</li>
        <li>Choisissez <strong>S'abonner depuis le web</strong> (ou <strong>Depuis Internet</strong> dans l'application de bureau).</li>
        <li>Collez l'URL indiquée ci-dessus, puis cliquez sur <strong>Importer</strong> ou <strong>OK</strong>.</li>
      </ol>
      </div>
    </div>

    <div id="tab-apple" class="tab-content steps">
      <ol>
        <li>Open the <strong>Calendar</strong> app on your Mac.</li>
        <li>From the menu bar, choose <strong>File &gt; New Calendar Subscription</strong>.</li>
        <li>Paste the calendar URL shown above, then click <strong>Subscribe</strong>.</li>
        <li>Give it a name, choose a colour, and set <strong>Auto-refresh</strong> to <strong>Every day</strong> or <strong>Every week</strong>.</li>
        <li>Click <strong>OK</strong>.</li>
      </ol>
      </div>
      <div data-fr>
      <ol>
        <li>Ouvrez l'application <strong>Calendrier</strong> sur votre Mac.</li>
        <li>Dans la barre de menus, choisissez <strong>Fichier &gt; Nouvel abonnement à un calendrier</strong>.</li>
        <li>Collez l'URL indiquée ci-dessus, puis cliquez sur <strong>S'abonner</strong>.</li>
        <li>Donnez-lui un nom, choisissez une couleur et définissez <strong>Actualisation automatique</strong> sur <strong>Chaque jour</strong> ou <strong>Chaque semaine</strong>.</li>
        <li>Cliquez sur <strong>OK</strong>.</li>
      </ol>
      </div>
    </div>
  </div>

  <div class="filters">
    <button class="filter-btn active" onclick="filter('all',this)"><span data-en>All acts</span><span data-fr>Tous les groupes</span></button>
    <button class="filter-btn" onclick="filter('Thu 30 Jul',this)">Jeu 30 Jui</button>
    <button class="filter-btn" onclick="filter('Fri 31 Jul',this)">Ven 31 Jui</button>
    <button class="filter-btn" onclick="filter('Sat 1 Aug',this)">Sam 1 Aoû</button>
    <button class="filter-btn" onclick="filter('Sun 2 Aug',this)">Dim 2 Aoû</button>
    <button class="filter-btn" onclick="filter('Main Stage',this)"><span data-en>Main Stage</span><span data-fr>Grande Scène</span></button>
    <button class="filter-btn" onclick="filter('Day Stage',this)"><span data-en>Day Stage</span><span data-fr>Scène de Jour</span></button>
  </div>
  <div id="band-list"></div>
</main>
<footer>
  <p>
    <a href="https://www.larochebluegrass.org" target="_blank">larochebluegrass.org</a>
    &nbsp;&middot;&nbsp;
    <span data-en>Concerts calendar:</span><span data-fr>Calendrier concerts :</span> <a href="https://tinyurl.com/LaRoche2026">tinyurl.com/LaRoche2026</a>
    &nbsp;&middot;&nbsp;
    <span data-en>Updated automatically from the festival schedule</span><span data-fr>Mis à jour automatiquement depuis le programme officiel</span>
  </p>
  <p style="margin-top:0.5rem;">&copy; 2026 Ken Gray &nbsp;&middot;&nbsp; Poster &copy; Paul Boutet / Roch&apos;&eacute;v&eacute;nements &nbsp;&middot;&nbsp; <span data-en>Built with</span><span data-fr>Réalisé avec</span> <a href="https://claude.ai" target="_blank">Claude</a></p>
</footer>
<script>
const CAL_URLS = {{
  festival: 'https://tinyurl.com/LaRoche2026-Festival',
  camp: 'https://tinyurl.com/LaRoche2026-Camp',
  full: 'https://tinyurl.com/LaRoche2026-Full'
}};
let currentCalUrl = CAL_URLS.festival;
let currentLang = 'en';
function setCalType(type, btn) {{
  currentCalUrl = CAL_URLS[type];
  document.getElementById('cal-url-display').childNodes[0].textContent = currentCalUrl + ' ';
  document.querySelectorAll('.platform-tabs:first-of-type .tab-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
}}
function setLang(lang, btn) {{
  currentLang = lang;
  document.getElementById('html-root').className = lang === 'fr' ? 'fr' : '';
  document.querySelectorAll('.lang-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  render(currentFilter);
}}
const bands = {bands_js};
const dayNames = {{
  en: {{'Thu 30 Jul':'Thursday 30 July','Fri 31 Jul':'Friday 31 July','Sat 1 Aug':'Saturday 1 August','Sun 2 Aug':'Sunday 2 August'}},
  fr: {{'Thu 30 Jul':'Jeudi 30 juillet','Fri 31 Jul':'Vendredi 31 juillet','Sat 1 Aug':'Samedi 1er août','Sun 2 Aug':'Dimanche 2 août'}}
}};
const days = ['Thu 30 Jul','Fri 31 Jul','Sat 1 Aug','Sun 2 Aug'];
let currentFilter = 'all';
function pillClass(s) {{ return s === 'Main Stage' ? 'pill-main' : 'pill-day'; }}
function pillLabel(s) {{
  if (currentLang === 'fr') return s === 'Main Stage' ? 'Grande Scène' : 'Scène de Jour';
  return s;
}}
function render(filterBy) {{
  currentFilter = filterBy;
  const container = document.getElementById('band-list');
  let html = '';
  days.forEach(day => {{
    const dayBands = bands.filter(b => {{
      if (filterBy === 'all') return b.date === day;
      if (filterBy === day)   return b.date === day;
      if (filterBy === 'Main Stage' || filterBy === 'Day Stage') return b.date === day && b.stage === filterBy;
      return false;
    }}).sort((a, b) => a.time.localeCompare(b.time));
    if (!dayBands.length) return;
    html += `<div class="day-section"><div class="day-header"><span class="day-title">${{dayNames[currentLang][day]}}</span><span class="day-count">${{dayBands.length}} ${{currentLang === 'fr' ? (dayBands.length !== 1 ? 'groupes' : 'groupe') : (dayBands.length !== 1 ? 'acts' : 'act')}}</span></div><div class="band-grid">`;
    dayBands.forEach(b => {{
      const links = [];
      if (b.website) links.push(`<a class="card-link" href="${{b.website}}" target="_blank" onclick="event.stopPropagation()">Website</a>`);
      if (b.video)   links.push(`<a class="card-link" href="${{b.video}}" target="_blank" onclick="event.stopPropagation()">Video</a>`);
      if (b.spotify) links.push(`<a class="card-link spotify" href="${{b.spotify}}" target="_blank" onclick="event.stopPropagation()">Spotify</a>`);
      const LIMIT = 180;
      const noteText = currentLang === 'fr' ? (b.notes_fr || b.notes) : b.notes;
      const notesHtml = noteText
        ? (noteText.length > LIMIT
            ? `<p class="card-notes">${{noteText.slice(0, LIMIT).trimEnd()}}<span class="expand-ellipsis" onclick="expandCard(event, this)">...</span><span class="notes-rest" style="display:none">${{noteText.slice(LIMIT)}}</span></p>`
            : `<p class="card-notes">${{noteText}}</p>`)
        : '';
      html += `<div class="band-card"><div class="card-header"><span class="band-name">${{b.name}}</span><span class="stage-pill ${{pillClass(b.stage)}}">${{pillLabel(b.stage)}}</span></div><div class="card-meta">${{b.time}}${{b.country ? ' &middot; ' + b.country : ''}}</div>${{notesHtml}}<div class="card-links">${{links.join('')}}</div></div>`;
    }});
    html += '</div></div>';
  }});
  container.innerHTML = html || '<p style="color:var(--text-muted);padding:1rem 0;">No acts found.</p>';
}}
function filter(value, btn) {{
  document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  render(value);
}}
function expandCard(e, el) {{
  e.stopPropagation();
  const rest = el.nextElementSibling;
  rest.style.display = 'inline';
  el.remove();
}}
function toggleInstructions(e) {{
  e.preventDefault();
  const section = document.getElementById('subscribe-section');
  section.style.display = section.style.display === 'none' ? 'block' : 'none';
}}
function showTab(id, btn) {{
  document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('tab-' + id).classList.add('active');
  btn.classList.add('active');
}}
function copyUrl() {{
  navigator.clipboard.writeText(currentCalUrl).then(() => {{
    const btns = document.querySelectorAll('.copy-btn');
    btns.forEach(b => {{ b.textContent = 'Copied!'; setTimeout(() => b.textContent = 'Copy', 2000); }});
  }});
}}
render('all');
</script>
</body>
</html>'''

html_path = repo_root / "index.html"
html_path.write_text(html_content, encoding="utf-8")
print(f"HTML written: {html_path}")

