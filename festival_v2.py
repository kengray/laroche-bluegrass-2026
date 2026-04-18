# =============================================================================
# festival_v2.py
# Generates the Bluegrass in La Roche 2026 festival schedule spreadsheet.
#
# Requires: openpyxl  (pip install openpyxl)
# Output:   LaRoche2026_Festival_Schedule.xlsx
#
# The workbook contains five sheets:
#   1. Main Stage     — evening concerts, Thu 30 Jul – Sun 2 Aug
#   2. Day Stage      — lunchtime concerts, Sat 1 Aug & Sun 2 Aug
#   3. Street Festival — free outdoor concerts (TBC placeholders)
#   4. Teaching Camp  — full timetable for the 3-day teaching camp
#   5. Google Calendar Import — pre-formatted for CSV export & import
# =============================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new empty workbook. The first sheet is created automatically
# and is accessed via wb.active; subsequent sheets use wb.create_sheet().
wb = Workbook()

# -----------------------------------------------------------------------------
# Colour palette — all colours are hex RGB strings without the leading #.
# Each stage has a dark header colour and a light row-highlight colour.
# -----------------------------------------------------------------------------
DARK_GREEN  = "1A3C34"   # Title bar background
MID_GREEN   = "2D6A4F"   # Main Stage tab and header
LIGHT_GREEN = "D8F3DC"   # Main Stage date-change highlight row
GOLD        = "B7950B"   # Day Stage tab, header, and time column text
GOLD_LIGHT  = "FFF9C4"   # Day Stage date-change highlight row
WHITE       = "FFFFFF"   # Alternating row background (odd)
GREY_ROW    = "F4F4F4"   # Alternating row background (even)
DARK_RED    = "8B0000"   # Google Calendar Import tab and header
STREET_BLUE = "1A3A5C"   # Street Festival tab and header
STREET_LIGHT= "D6E4F0"   # Street Festival date-change highlight row
CAMP_PURPLE = "4A235A"   # Teaching Camp tab and header
CAMP_LIGHT  = "EAD5F5"   # Teaching Camp date-change highlight row


def thin_border():
    """
    Returns a Border object with a thin light-grey line on all four sides.
    Called for every data cell to give the table a clean grid appearance.
    """
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def make_sheet(wb, title, rows, tab_color, header_color, light_row_color, is_first=False):
    """
    Builds a single formatted worksheet and populates it with schedule data.

    Parameters
    ----------
    wb              : openpyxl Workbook — the workbook to add the sheet to
    title           : str  — the sheet tab name (e.g. "Main Stage")
    rows            : list — list of data rows; each row is a list:
                      [date, name, country/location, stage, start, end, notes]
    tab_color       : str  — hex colour for the sheet tab
    header_color    : str  — hex colour for the column header row background
    light_row_color : str  — hex colour used when the date changes (visual grouping)
    is_first        : bool — if True, uses the default active sheet instead of
                      creating a new one (avoids an unwanted blank "Sheet" tab)

    Layout
    ------
    Row 1 : merged title banner
    Row 2 : merged subtitle / instructions
    Row 3 : column headers
    Row 4 : greyed-out example row showing expected time format
    Row 5+ : data rows
    """

    # Use the pre-existing active sheet for the first tab; create new ones for the rest
    ws = wb.active if is_first else wb.create_sheet(title)
    ws.title = title
    ws.sheet_view.showGridLines = False   # Hide the default Excel gridlines
    ws.sheet_properties.tabColor = tab_color

    # Column headers and their widths (in Excel character units)
    headers    = ["Date", "Band / Act", "Country", "Stage Note", "Start Time", "End Time", "Notes"]
    col_widths = [14,      36,           22,         20,           13,           13,          45]

    # --- Row 1: Title banner ---------------------------------------------------
    ws.merge_cells("A1:G1")   # Span all 7 columns
    title_cell = ws["A1"]
    title_cell.value     = f"Bluegrass in La Roche 2026 — {title}"
    title_cell.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    title_cell.fill      = PatternFill("solid", fgColor=DARK_GREEN)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # --- Row 2: Subtitle / instructions ---------------------------------------
    ws.merge_cells("A2:G2")
    sub           = ws["A2"]
    sub.value     = "Fill in Start Time and End Time after consulting the schedule image on the festival website"
    sub.font      = Font(name="Arial", italic=True, size=9, color="666666")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # --- Row 3: Column headers ------------------------------------------------
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell           = ws.cell(row=3, column=col, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20

    # --- Row 4: Example row ---------------------------------------------------
    # Shown in light grey italic to demonstrate the expected 24hr time format
    example = ["e.g. Thu 30 Jul", "e.g. Kristy Cox Band", "AU/US", "Main Stage",
               "19:00", "20:00", "Times in 24hr format — e.g. 19:00 to 20:00"]
    for col, val in enumerate(example, 1):
        cell           = ws.cell(row=4, column=col, value=val)
        cell.font      = Font(name="Arial", size=9, italic=True, color="AAAAAA")
        cell.fill      = PatternFill("solid", fgColor="F0F0F0")
        cell.border    = thin_border()
        # Centre the time columns; left-align everything else
        cell.alignment = Alignment(
            vertical="center",
            horizontal="center" if col in (5, 6) else "left",
            wrap_text=True
        )
    ws.row_dimensions[4].height = 24

    # --- Rows 5+: Data rows ---------------------------------------------------
    current_date = None   # Tracks date changes to apply the highlight colour

    for i, row in enumerate(rows, start=5):
        is_tbc = row[1].startswith("TBC")   # True for Street Festival placeholders

        # Apply light_row_color when the date changes (visual day grouping),
        # then alternate white/grey within the same day
        if row[0] != current_date:
            current_date = row[0]
            bg = light_row_color
        else:
            bg = GREY_ROW if i % 2 == 0 else WHITE

        # Override with amber for TBC placeholder rows so they stand out
        if is_tbc:
            bg = "FFF3CD"

        for col, val in enumerate(row, 1):
            cell        = ws.cell(row=i, column=col, value=val)
            cell.fill   = PatternFill("solid", fgColor=bg)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # TBC rows: grey italic text; normal rows: black text
            cell.font = Font(
                name="Arial", size=9,
                italic=is_tbc,
                color="888888" if is_tbc else "000000"
            )

            # Time columns (Start Time = col 5, End Time = col 6):
            # Centre-align and use gold bold text for confirmed entries
            if col in (5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(
                    name="Arial", size=9,
                    color=GOLD,
                    bold=not is_tbc,
                    italic=is_tbc
                )

        ws.row_dimensions[i].height = 30

    # Freeze rows 1-4 so headers stay visible when scrolling
    ws.freeze_panes = "A5"
    return ws


# =============================================================================
# SCHEDULE DATA
# Each row is a list: [date, name, country, stage, start_time, end_time, notes]
# Times are in 24hr format. End times past midnight roll correctly in the ICS
# generator. Source: official schedule image from larochebluegrass.org
# =============================================================================

# -----------------------------------------------------------------------------
# MAIN STAGE — evening concerts from 18:00 (17:30 on Thursday)
# Each act ends when the next one starts; closing acts have estimated end times
# -----------------------------------------------------------------------------
main_rows = [
    # Thursday 30 July
    ["Thu 30 Jul", "Workshop Students & Teachers", "Europe",         "Main Stage", "17:30", "20:00", "End of teaching camp — students take the stage"],
    ["Thu 30 Jul", "Piotr Bulas & Garage Folks",   "PL",             "Main Stage", "20:00", "21:30", "Traditional Bluegrass with humour; luthier-led band"],
    ["Thu 30 Jul", "Opening Ceremony",              "",               "Main Stage", "21:30", "22:00", ""],
    ["Thu 30 Jul", "Paul & The Alley Cats",         "NL/ES",          "Main Stage", "22:00", "23:30", "European pioneers; Paul Von Vlodrop on mandolin"],
    ["Thu 30 Jul", "Kristy Cox Band",               "AU/US",          "Main Stage", "23:30", "01:00", "Nashville-based; 5 studio albums; returning since 2014"],
    # Friday 31 July
    ["Fri 31 Jul", "Noah Hassler-Forest",           "NL/ES",          "Main Stage", "18:00", "19:00", "Rotterdam contest winners; jazz & world music-infused Bluegrass"],
    ["Fri 31 Jul", "No Man's Land",                 "SK/US/NL/DK/BE", "Main Stage", "19:00", "20:00", "All-female European supergroup; La Roche debut as full band"],
    ["Fri 31 Jul", "The New Aliquot & David Benda", "CZ",             "Main Stage", "20:00", "21:30", "Ondra Kozák; 4-time La Roche contest winners; 10th anniversary"],
    ["Fri 31 Jul", "Sentimental Gentlemen",         "US",             "Main Stage", "21:30", "23:00", "East Nashville; Sierra Ferrell's touring band"],
    ["Fri 31 Jul", "Veranda",                       "CA",             "Main Stage", "23:00", "00:30", "Montreal; Bluegrass in English and French"],
    # Saturday 1 August
    ["Sat 1 Aug",  "Yonder Boys",                   "DE/US/AU/CL",    "Main Stage", "18:00", "19:00", "Berlin-based; Americana/Bluegrass/folk; promoted from Day Stage"],
    ["Sat 1 Aug",  "Sweet Sally",                   "US",             "Main Stage", "19:00", "20:00", "Young California trio; Berklee scholars; CBA stable"],
    ["Sat 1 Aug",  "Country Gongbang",              "KO",             "Main Stage", "20:00", "21:30", "Korean Bluegrass; IBMA grant recipients; Grand Ole Opry veterans"],
    ["Sat 1 Aug",  "Shelby Means Band",             "US",             "Main Stage", "21:30", "23:00", "2x Grammy winner (ex-Molly Tuttle); solo album 2025"],
    ["Sat 1 Aug",  "Monogram",                      "CZ",             "Main Stage", "23:00", "00:30", "30-year-old band; La Roche contest winners 2008; first visit in 10 years"],
    # Sunday 2 August
    ["Sun 2 Aug",  "William Jack",                  "AU",             "Main Stage", "18:00", "19:00", "Solo cello; classically trained; Chris Thile & Tony Rice influences"],
    ["Sun 2 Aug",  "Blue Lass",                     "GB",             "Main Stage", "19:00", "20:00", "Promoted from Day Stage 2024; Bluegrass/Old Time/Folk mix"],
    ["Sun 2 Aug",  "Gadan",                         "IT/IE/US",       "Main Stage", "20:00", "21:15", "Celtic-Americana fusion; 3 banjo players incl. Enda Scahill (We Banjo Three)"],
    ["Sun 2 Aug",  "Hayde Bluegrass Orchestra",     "NO",             "Main Stage", "21:15", "22:30", "2x IBMA Momentum nominees 2022; touring US regulars"],
    ["Sun 2 Aug",  "Yellofox",                      "NL",             "Main Stage", "22:30", "00:00", "Folk-Americana; Paul Simon/Gram Parsons influences; festival closer"],
]

# -----------------------------------------------------------------------------
# DAY STAGE — lunchtime concerts 12:00-17:00, Saturday & Sunday only
# Each slot is exactly 1 hour as shown in the official schedule image
# -----------------------------------------------------------------------------
day_rows = [
    # Saturday 1 August
    ["Sat 1 Aug", "Kids on Bluegrass Europe",       "Europe",  "Day Stage", "12:00", "13:00", "Ages 6-16; two mornings of workshops culminating in performance"],
    ["Sat 1 Aug", "Kristy Cox Band",                "AU/US",   "Day Stage", "13:00", "14:00", "Also headlining Thu Main Stage; second chance to see them"],
    ["Sat 1 Aug", "That's All Folk",                "FR",      "Day Stage", "14:00", "15:00", "Local Alps trio; traditional Bluegrass + modern pop covers"],
    ["Sat 1 Aug", "The Sentimental Gentlemen",      "US",      "Day Stage", "15:00", "16:00", "Also on Main Stage Fri; Sierra Ferrell's touring band"],
    ["Sat 1 Aug", "Tim O'Connor Trio",              "IE/FR",   "Day Stage", "16:00", "17:00", "Irishman in Aix-les-Bains; folk/Irish/blues/rock trio"],
    # Sunday 2 August
    ["Sun 2 Aug", "Grace Honeywell & Jeri Foreman", "US/AU",   "Day Stage", "12:00", "13:00", "Twin fiddles; folk/country; formed in UK"],
    ["Sun 2 Aug", "Catch the Goat",                 "FR",      "Day Stage", "13:00", "14:00", "Toulouse veterans; Dawgrass style; mandocello & cello"],
    ["Sun 2 Aug", "Shelby Means Band",              "US",      "Day Stage", "14:00", "15:00", "Also on Sat Main Stage; entirely different set"],
    ["Sun 2 Aug", "Country Gongbang",               "KO",      "Day Stage", "15:00", "16:00", "Also on Sat Main Stage; final chance before they fly home"],
    ["Sun 2 Aug", "Sweet Sally",                    "US",      "Day Stage", "16:00", "17:00", "Also on Sat Main Stage; California trio"],
]

# -----------------------------------------------------------------------------
# STREET FESTIVAL — free concerts around town
# Wednesday evening & Friday lunchtime. ~12 terrace concerts + 3 offsite.
# Full programme not yet published; TBC placeholder rows included so the
# structure is ready to fill in when the festival announces the line-up.
# -----------------------------------------------------------------------------
street_rows = [
    # Wednesday 29 July — evening
    ["Wed 29 Jul", "TBC — Street Act 1",      "TBC", "Street Festival",           "", "", "Bar/restaurant/café terraces around town centre. FREE. Band TBC — check festival website."],
    ["Wed 29 Jul", "TBC — Street Act 2",      "TBC", "Street Festival",           "", "", "Bar/restaurant/café terraces around town centre. FREE. Band TBC — check festival website."],
    ["Wed 29 Jul", "TBC — Street Act 3",      "TBC", "Street Festival",           "", "", "Bar/restaurant/café terraces around town centre. FREE. Band TBC — check festival website."],
    ["Wed 29 Jul", "TBC — Street Act 4",      "TBC", "Street Festival",           "", "", "Bar/restaurant/café terraces around town centre. FREE. Band TBC — check festival website."],
    ["Wed 29 Jul", "TBC — Street Act 5",      "TBC", "Street Festival",           "", "", "Bar/restaurant/café terraces around town centre. FREE. Band TBC — check festival website."],
    ["Wed 29 Jul", "TBC — Street Act 6",      "TBC", "Street Festival",           "", "", "Bar/restaurant/café terraces around town centre. FREE. Band TBC — check festival website."],
    ["Wed 29 Jul", "TBC — Offsite Concert 1", "TBC", "Street Festival (Offsite)", "", "", "Health centre concert. FREE. Venue & band TBC — check festival website."],
    # Friday 31 July — lunchtime
    ["Fri 31 Jul", "TBC — Street Act 7",      "TBC", "Street Festival",           "", "", "Bar/restaurant/café terraces around town centre. FREE. Band TBC — check festival website."],
    ["Fri 31 Jul", "TBC — Street Act 8",      "TBC", "Street Festival",           "", "", "Bar/restaurant/café terraces around town centre. FREE. Band TBC — check festival website."],
    ["Fri 31 Jul", "TBC — Street Act 9",      "TBC", "Street Festival",           "", "", "Bar/restaurant/café terraces around town centre. FREE. Band TBC — check festival website."],
    ["Fri 31 Jul", "TBC — Street Act 10",     "TBC", "Street Festival",           "", "", "Bar/restaurant/café terraces around town centre. FREE. Band TBC — check festival website."],
    ["Fri 31 Jul", "TBC — Street Act 11",     "TBC", "Street Festival",           "", "", "Bar/restaurant/café terraces around town centre. FREE. Band TBC — check festival website."],
    ["Fri 31 Jul", "TBC — Street Act 12",     "TBC", "Street Festival",           "", "", "Bar/restaurant/café terraces around town centre. FREE. Band TBC — check festival website."],
    ["Fri 31 Jul", "TBC — Offsite Concert 2", "TBC", "Street Festival (Offsite)", "", "", "Health centre concert. FREE. Venue & band TBC — check festival website."],
    ["Fri 31 Jul", "TBC — Offsite Concert 3", "TBC", "Street Festival (Offsite)", "", "", "Health centre concert. FREE. Venue & band TBC — check festival website."],
]

# Build the first three sheets
make_sheet(wb, "Main Stage",      main_rows,   MID_GREEN,   MID_GREEN,   LIGHT_GREEN, is_first=True)
make_sheet(wb, "Day Stage",       day_rows,    GOLD,        GOLD,        GOLD_LIGHT,  is_first=False)
make_sheet(wb, "Street Festival", street_rows, STREET_BLUE, STREET_BLUE, STREET_LIGHT,is_first=False)

# -----------------------------------------------------------------------------
# TEACHING CAMP — 3-day residential camp at Lycée Sainte Famille
# Monday 27 July (arrival) through Thursday 30 July (students perform on stage)
# Timetable source: official camp timetable image on larochebluegrass.org
# Note: the "Stage Note" column holds the location within the camp venue
# -----------------------------------------------------------------------------
camp_rows = [
    # Monday 27 July — arrival day
    ["Mon 27 Jul", "Accommodation check-in",              "Lycée Sainte Famille", "14:00", "17:30", "School open for registration and room check-in"],
    ["Mon 27 Jul", "Welcome talk & drinks",               "Lycée Sainte Famille", "17:30", "20:00", "Welcome by Festival Chair and Camp Leader Gilles Rézard"],
    ["Mon 27 Jul", "Dinner",                              "Lycée Sainte Famille", "20:00", "21:30", ""],
    # Tuesday 28 July — first full day
    ["Tue 28 Jul", "Breakfast",                           "Lycée Sainte Famille", "07:45", "09:00", ""],
    ["Tue 28 Jul", "Today's news + Instrumental classes", "Lycée Sainte Famille", "09:00", "10:30", "Morning session 1 — instrumental classes"],
    ["Tue 28 Jul", "Break",                               "Lycée Sainte Famille", "10:30", "11:00", ""],
    ["Tue 28 Jul", "Instrumental classes",                "Lycée Sainte Famille", "11:00", "12:30", "Morning session 2 — instrumental classes"],
    ["Tue 28 Jul", "Lunch",                               "Lycée Sainte Famille", "12:30", "14:00", ""],
    ["Tue 28 Jul", "Wernick Method Jam: Instructions",    "Lycée Sainte Famille", "14:00", "16:00", "Afternoon — Wernick Method jam instructions"],
    ["Tue 28 Jul", "Break",                               "Lycée Sainte Famille", "16:00", "16:30", ""],
    ["Tue 28 Jul", "Optional/requested modules",          "Lycée Sainte Famille", "16:30", "19:00", "Requested/optional modules"],
    ["Tue 28 Jul", "Dinner",                              "Lycée Sainte Famille", "19:00", "20:30", ""],
    # Wednesday 29 July — second full day
    ["Wed 29 Jul", "Breakfast",                           "Lycée Sainte Famille", "07:45", "09:00", ""],
    ["Wed 29 Jul", "Today's news + Instrumental classes", "Lycée Sainte Famille", "09:00", "10:30", "Morning session 1 — instrumental classes"],
    ["Wed 29 Jul", "Break",                               "Lycée Sainte Famille", "10:30", "11:00", ""],
    ["Wed 29 Jul", "Instrumental classes",                "Lycée Sainte Famille", "11:00", "12:30", "Morning session 2 — instrumental classes"],
    ["Wed 29 Jul", "Lunch",                               "Lycée Sainte Famille", "12:30", "14:00", ""],
    ["Wed 29 Jul", "Wernick Method Jam: Coached jam",     "Lycée Sainte Famille", "14:00", "16:00", "Afternoon — coached Wernick Method jam"],
    ["Wed 29 Jul", "Break",                               "Lycée Sainte Famille", "16:00", "16:30", ""],
    ["Wed 29 Jul", "Concert rehearsal & organisation",    "Lycée Sainte Famille", "16:30", "18:00", "Rehearsal for Thursday student concert"],
    ["Wed 29 Jul", "Dinner",                              "Lycée Sainte Famille", "18:00", "19:00", ""],
    # Thursday 30 July — final camp day; students perform at festival in evening
    ["Thu 30 Jul", "Breakfast",                           "Lycée Sainte Famille", "07:45", "09:00", ""],
    ["Thu 30 Jul", "Today's news + Instrumental classes", "Lycée Sainte Famille", "09:00", "10:30", "Morning session 1 — instrumental classes"],
    ["Thu 30 Jul", "Break",                               "Lycée Sainte Famille", "10:30", "11:00", ""],
    ["Thu 30 Jul", "Band labs / jams",                    "Lycée Sainte Famille", "11:00", "12:30", "Final morning session — band labs and jams"],
    ["Thu 30 Jul", "Lunch",                               "Lycée Sainte Famille", "12:30", "14:00", ""],
    ["Thu 30 Jul", "Free time",                           "Lycée Sainte Famille", "14:00", "16:00", ""],
    ["Thu 30 Jul", "Break",                               "Lycée Sainte Famille", "16:00", "16:30", ""],
    ["Thu 30 Jul", "Head to festival site",               "Lycée Sainte Famille", "16:30", "17:30", "Move to Lycée Sainte Marie festival site"],
    ["Thu 30 Jul", "Students concert — Main Stage",       "Festival Main Stage",  "17:30", "19:00", "Band Labs perform on the Main Stage"],
    ["Thu 30 Jul", "French teachers set — Main Stage",    "Festival Main Stage",  "19:00", "20:00", "French teachers perform on the Main Stage"],
]
make_sheet(wb, "Teaching Camp", camp_rows, CAMP_PURPLE, CAMP_PURPLE, CAMP_LIGHT, is_first=False)


# =============================================================================
# GOOGLE CALENDAR IMPORT SHEET
# Pre-formatted with Subject, Start Date, Start Time, End Date, End Time,
# Description columns. Once times are filled in, export this sheet as CSV
# and import into Google Calendar via File > Import.
# =============================================================================
gc = wb.create_sheet("Google Calendar Import")
gc.sheet_properties.tabColor = DARK_RED
gc.sheet_view.showGridLines = False

# Title banner
gc.merge_cells("A1:F1")
t           = gc["A1"]
t.value     = "Google Calendar Import — fill in times, then export this sheet as CSV and import into Google Calendar"
t.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
t.fill      = PatternFill("solid", fgColor=DARK_RED)
t.alignment = Alignment(horizontal="center", vertical="center")
gc.row_dimensions[1].height = 26

# Subtitle note
gc.merge_cells("A2:F2")
note           = gc["A2"]
note.value     = "TBC rows are placeholders for the Street Festival — update when the festival publishes the full street programme"
note.font      = Font(name="Arial", italic=True, size=9, color="888888")
note.alignment = Alignment(horizontal="center")
gc.row_dimensions[2].height = 16

# Column headers — matches the format Google Calendar expects for CSV import
gc_headers = ["Subject", "Start Date", "Start Time", "End Date", "End Time", "Description"]
gc_widths   = [50,        14,           13,            14,          13,          60]
for col, (h, w) in enumerate(zip(gc_headers, gc_widths), 1):
    cell           = gc.cell(row=3, column=col, value=h)
    cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    cell.fill      = PatternFill("solid", fgColor=DARK_RED)
    cell.alignment = Alignment(horizontal="center")
    cell.border    = thin_border()
    gc.column_dimensions[get_column_letter(col)].width = w

# Lookup table: display date string -> DD/MM/YYYY format for Google Calendar
date_map = {
    "Mon 27 Jul": "27/07/2026",
    "Tue 28 Jul": "28/07/2026",
    "Wed 29 Jul": "29/07/2026",
    "Thu 30 Jul": "30/07/2026",
    "Fri 31 Jul": "31/07/2026",
    "Sat 1 Aug":  "01/08/2026",
    "Sun 2 Aug":  "02/08/2026",
}

# Combine all data rows from all four sections, tagging each with its stage name
all_rows = (
    [(r, "Main Stage")      for r in main_rows]   +
    [(r, "Day Stage")       for r in day_rows]    +
    [(r, "Street Festival") for r in street_rows] +
    [(r, "Teaching Camp")   for r in camp_rows]
)

# Write one row per event into the Google Calendar Import sheet
for i, (row, stage) in enumerate(all_rows, start=4):
    date_str = date_map.get(row[0], row[0])
    is_tbc   = row[1].startswith("TBC")

    # Build the event subject line: "Band name (Country) — Stage"
    subject = f"{row[1]} ({row[2]}) — {stage}"
    # Description is the last element; guard against rows that don't have one
    desc    = row[6] if len(row) > 6 else ""

    # Subject, Start Date, Start Time, End Date, End Time, Description
    vals = [subject, date_str, row[4], date_str, row[5], desc]

    # Amber background for TBC rows; alternating white/grey for confirmed rows
    bg = "FFF3CD" if is_tbc else (GREY_ROW if i % 2 == 0 else WHITE)

    for col, val in enumerate(vals, 1):
        cell           = gc.cell(row=i, column=col, value=val)
        cell.font      = Font(name="Arial", size=9, italic=is_tbc,
                              color="888888" if is_tbc else "000000")
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.border    = thin_border()
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    gc.row_dimensions[i].height = 20

# Freeze the header rows so they stay visible when scrolling
gc.freeze_panes = "A3"

# Save the completed workbook
wb.save("/home/claude/LaRoche2026_Festival_Schedule.xlsx")
print("Done")
